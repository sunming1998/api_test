import openpyxl
import os

class DoExcel():
    def __init__(self, file_path ,sheet_name):
        # 将方法封装起来，调用的时候，file_path 填Excel名就行
        self.file_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))+ r'\test_data\{}'.format(file_path)
        self.sheet_name = sheet_name

    def read_data(self):
        # 打开Excel表格
        wb = openpyxl.load_workbook(self.file_path)
        # 打开表单
        sheet = wb[self.sheet_name]
        # 创建一个空列表,将取出来的测试用例放进去
        test_data = []
        # 循环读取用例数据，存入数据
        for i in range(2,sheet.max_row+1):
            # 创建一个空字典，把元素都添加进去
            test_dict = {}
            test_dict['case_id'] = sheet.cell(i, 1).value
            test_dict['case_name'] = sheet.cell(i, 2).value
            test_dict['method'] = sheet.cell(i, 3).value
            test_dict['url'] = sheet.cell(i, 4).value
            test_dict['headers'] = sheet.cell(i, 5).value
            test_dict['data'] = sheet.cell(i,6).value
            test_data.append(test_dict)
        return test_data

    # 写入excel数据,参数为行、列、值
    def write_data(self, row, column, value):
        # 先打开Excel表格
        workbook = openpyxl.load_workbook(self.file_path)
        # 打开表单
        sheet = workbook[self.sheet_name]
        sheet.cell(row=row, column=column).value = value
        workbook.save(self.file_path)



if __name__ == '__main__':
    a = DoExcel('test_demo.xlsx','case')
    print(a.read_data())








